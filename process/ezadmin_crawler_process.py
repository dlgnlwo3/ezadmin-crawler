if 1 == 1:
    import sys
    import os

    sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))


from selenium import webdriver
from dtos.gui_dto import GUIDto

from common.utils import global_log_append
from common.chrome import open_browser, get_chrome_driver
from common.selenium_activities import close_new_tabs, alert_ok_try
from common.account_file import AccountFile


from enums.store_column_enum import CommonStoreEnum, Cafe24Enum, ElevenStreetEnum
from enums.store_name_enum import StoreNameEnum

from features.convert_store_name import StoreNameConverter

from dtos.store_detail_dto import StoreDetailDto

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains
from selenium.webdriver.support.select import Select

from datetime import timedelta, datetime
import time

import pandas as pd
from openpyxl import load_workbook

import re


class EzadminCrawlerProcess:
    def __init__(self):
        open_browser()
        self.default_wait = 10
        self.driver: webdriver.Chrome = get_chrome_driver(is_headless=False, is_secret=False)
        self.driver.implicitly_wait(self.default_wait)
        self.driver.maximize_window()

    def setGuiDto(self, guiDto: GUIDto):
        self.guiDto = guiDto

    def setLogger(self, log_msg):
        self.log_msg = log_msg

    def get_dict_account(self):
        df_accounts = AccountFile(self.guiDto.account_file).df_account
        df_accounts = df_accounts.fillna("")
        dict_accounts = {}
        for index, row in df_accounts.iterrows():
            channel = str(row["채널명"])
            domain = str(row["도메인"])
            account_id = str(row["ID"])
            account_pw = str(row["PW"])
            url = str(row["URL"])
            dict_accounts[channel] = {"도메인": domain, "ID": account_id, "PW": account_pw, "URL": url}
        return dict_accounts

    def get_store_list(self):
        # header 옵션에 2를 넣어서 3행부터 읽기 시작합니다.
        # 반드시 3행에 쇼핑몰의 이름이 위치해야 합니다.
        df_stats: pd.DataFrame = pd.read_excel(
            self.guiDto.stats_file, sheet_name=self.guiDto.sheet_name, keep_default_na="", header=2
        )
        store_list = []
        for column in df_stats.columns:
            if column.find("Unnamed") <= -1 and column.find("\n") <= -1 and column.find("합계") <= -1:
                store_list.append(column)
        print(store_list)
        self.log_msg.emit(f"{store_list}가 발견되었습니다.")
        return store_list

    def get_store_min_col(self, store_name):
        merged_cells = self.sheet.merged_cells

        for merged_cell in merged_cells:
            if merged_cell.start_cell.internal_value == store_name:
                print(merged_cell.start_cell.internal_value)
                # store_column_range = merged_cell.coord
                store_min_col = merged_cell.min_col
                store_max_col = merged_cell.max_col
                store_column_range = [store_min_col, store_max_col]
                break

        # store_column_range = re.sub(r"\d+", "", store_column_range)
        print(store_column_range)
        return store_min_col

    def get_target_date_row(self, target_date):
        # 순회할 셀 범위 지정
        min_row, max_row = 1, self.sheet.max_row
        min_col, max_col = 1, self.sheet.max_column

        for row in self.sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if target_date in str(cell.value):
                    print(f"'{target_date}'이 포함된 셀 위치: ({cell.row}, {cell.column})")
                    return cell.row

    def ezadmin_login(self):
        driver = self.driver
        self.driver.get(self.dict_accounts["이지어드민"]["URL"])
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//body[@class="ezadmin-main-body"]'))
        )
        time.sleep(0.2)

        login_domain = self.dict_accounts["이지어드민"]["도메인"]
        login_id = self.dict_accounts["이지어드민"]["ID"]
        login_pw = self.dict_accounts["이지어드민"]["PW"]

        # 로그인 시도
        # 이 행위 중 하나라도 실패한다면 로그인 실패
        try:
            driver.implicitly_wait(2)

            open_button = driver.find_element(By.XPATH, '//a[./span[@class="img_login"]][contains(text(), "로그인")]')
            driver.execute_script("arguments[0].click();", open_button)
            time.sleep(0.2)

            domain_input = driver.find_element(By.CSS_SELECTOR, 'input[id="login-domain"]')
            domain_input.clear()
            time.sleep(0.2)
            domain_input.send_keys(login_domain)
            time.sleep(0.2)

            id_input = driver.find_element(By.CSS_SELECTOR, 'input[id="login-id"]')
            id_input.clear()
            time.sleep(0.2)
            id_input.send_keys(login_id)
            time.sleep(0.2)

            pwd_input = driver.find_element(By.CSS_SELECTOR, 'input[id="login-pwd"]')
            pwd_input.clear()
            time.sleep(0.2)
            pwd_input.send_keys(login_pw)
            time.sleep(0.2)

            save_domain = driver.find_element(By.XPATH, '//input[@id="savedomain"]')
            driver.execute_script("arguments[0].click();", save_domain)
            time.sleep(0.2)

            login_button = driver.find_element(By.XPATH, '//input[@class="login-btn" and @value="로그인"]')
            driver.execute_script("arguments[0].click();", login_button)
            time.sleep(0.2)

            # 로그인 성공 시 나오는 화면
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//body[@class="bgline"]')))
            time.sleep(0.2)

            self.close_ezadmin_notice_popups()

        except Exception as e:
            print(e)
            raise Exception(f"이지어드민 로그인 실패")

        finally:
            driver.implicitly_wait(self.default_wait)

    # 이지어드민 로그인 시 발생하는 팝업창을 모두 닫습니다.
    def close_ezadmin_notice_popups(self):
        driver = self.driver
        try:
            driver.implicitly_wait(1)

            driver.execute_script("hide_board('internal_board');")
            time.sleep(0.2)

            driver.execute_script("hide_board('sys_notice_board');")
            time.sleep(0.2)

            # $x('//a[contains(text(), "팝업 전체 닫기")]')
            close_all_popups = driver.find_element(By.XPATH, '//a[contains(text(), "팝업 전체 닫기")]')
            driver.execute_script("arguments[0].click();", close_all_popups)
            time.sleep(0.2)

        except Exception as e:
            print(e)

        finally:
            driver.implicitly_wait(self.default_wait)

    # 정산통계 -> 판매처별정산통계 화면으로 이동합니다.
    def go_store_calculate_menu_and_search_date(self):
        driver = self.driver
        driver.get("https://ga20.ezadmin.co.kr/template35.htm?template=F308")
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//h3[contains(text(), "판매처별정산통계")]'))
        )
        time.sleep(0.1)

        # 날짜 검색 타입
        query_type_select = Select(driver.find_element(By.CSS_SELECTOR, 'select[id="query_type"]'))
        query_type_select.select_by_visible_text("주문일")
        time.sleep(0.2)

        # 시작일
        start_date_input = driver.find_element(By.CSS_SELECTOR, 'input[id="start_date"]')
        start_date_input.clear()
        start_date_input.send_keys(self.guiDto.target_date)

        # 종료일
        end_date_input = driver.find_element(By.CSS_SELECTOR, 'input[id="end_date"]')
        end_date_input.clear()
        end_date_input.send_keys(self.guiDto.target_date)

        search_button = driver.find_element(By.XPATH, '//div[contains(@id, "search")][contains(text(), "검색")]')
        driver.execute_script("arguments[0].click();", search_button)
        time.sleep(3)

    def get_calculate_from_result(self, store_name: str, store_detail_dto: StoreDetailDto):
        driver = self.driver
        store_name = StoreNameConverter().convert_store_name(store_name)
        print(store_name)
        time.sleep(0.2)

        store_detail_dto.store_name = store_name

        # $x('//tr[./td[contains(text(), "11번가") and @class="shop_name"]]')
        try:
            result_tr = driver.find_element(
                By.XPATH, f'//tr[./td[contains(text(), "{store_name}") and @class="shop_name"]]'
            )

            # 주문수량
            tot_products = result_tr.find_element(
                By.CSS_SELECTOR, 'td[aria-describedby*="tot_products"]'
            ).get_attribute("textContent")
            store_detail_dto.tot_products = tot_products

            # 주문금액
            tot_amount = result_tr.find_element(By.CSS_SELECTOR, 'td[aria-describedby*="tot_amount"]').get_attribute(
                "textContent"
            )
            store_detail_dto.tot_amount = tot_amount

            # 상품원가
            org_price = result_tr.find_element(By.CSS_SELECTOR, 'td[aria-describedby*="org_price"]').get_attribute(
                "textContent"
            )
            store_detail_dto.org_price = org_price

        except Exception as e:
            print(f"{store_name} 검색 결과를 발견하지 못했습니다.")

        return store_detail_dto

    # 주문배송관리 -> 확장주문검색2 이동
    def go_store_cancel_menu_and_search_date(self, store_name: str, order_state: str):
        driver = self.driver
        driver.get("https://ga20.ezadmin.co.kr/template35.htm?template=DS00")
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//h3[contains(text(), "확장주문검색2")]')))
        time.sleep(0.1)

        # 날짜 검색 타입
        date_type_select = Select(driver.find_element(By.CSS_SELECTOR, 'select[name="date_type"]'))
        date_type_select.select_by_value("cancel_date")
        time.sleep(0.2)

        # 시작일
        start_date_input = driver.find_element(By.CSS_SELECTOR, 'input[id="start_date"]')
        start_date_input.clear()
        start_date_input.send_keys(self.guiDto.target_date)

        # 종료일
        end_date_input = driver.find_element(By.CSS_SELECTOR, 'input[id="end_date"]')
        end_date_input.clear()
        end_date_input.send_keys(self.guiDto.target_date)

        # C/S
        cs_select = Select(driver.find_element(By.CSS_SELECTOR, 'select[name="order_cs_sel"]'))
        if order_state == "취소":
            cs_select.select_by_visible_text("배송전 취소")
        elif order_state == "반품":
            cs_select.select_by_visible_text("배송후 취소")

        # 판매처
        store_select = Select(driver.find_element(By.CSS_SELECTOR, 'select[id="str_shop_code"]'))
        store_select.select_by_visible_text(store_name)
        time.sleep(0.2)

        search_button = driver.find_element(By.XPATH, '//div[contains(@id, "search")][contains(text(), "검색")]')
        driver.execute_script("arguments[0].click();", search_button)
        time.sleep(3)

    def get_cancel_from_result(self, store_detail_dto: StoreDetailDto):
        driver = self.driver
        time.sleep(0.2)

        try:
            # 상품수량
            cancel_total_data_product_sum = driver.find_element(
                By.CSS_SELECTOR, 'span[id="total_data_product_sum"]'
            ).get_attribute("textContent")
            store_detail_dto.cancel_total_data_product_sum = cancel_total_data_product_sum

            # 판매금액
            cancel_total_data_order_sum_amount = driver.find_element(
                By.CSS_SELECTOR, 'span[id="total_data_order_sum_amount"]'
            ).get_attribute("textContent")
            store_detail_dto.cancel_total_data_order_sum_amount = cancel_total_data_order_sum_amount

        except Exception as e:
            print(f"검색 결과를 발견하지 못했습니다.")

        return store_detail_dto

    def get_refund_from_result(self, store_detail_dto: StoreDetailDto):
        driver = self.driver
        time.sleep(0.2)

        try:
            # 상품수량
            refund_total_data_product_sum = driver.find_element(
                By.CSS_SELECTOR, 'span[id="total_data_product_sum"]'
            ).get_attribute("textContent")
            store_detail_dto.refund_total_data_product_sum = refund_total_data_product_sum

            # 판매금액
            refund_total_data_order_sum_amount = driver.find_element(
                By.CSS_SELECTOR, 'span[id="total_data_order_sum_amount"]'
            ).get_attribute("textContent")
            store_detail_dto.refund_total_data_order_sum_amount = refund_total_data_order_sum_amount

        except Exception as e:
            print(f"검색 결과를 발견하지 못했습니다.")

        return store_detail_dto

    # 정산통계 -> 배송통계 이동
    def go_store_delivery_menu_and_search_date(self, store_name: str):
        driver = self.driver
        driver.get("https://ga20.ezadmin.co.kr/template35.htm?template=F500")
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//h3[contains(text(), "배송통계")]')))
        time.sleep(0.1)

        # 판매처
        store_select = Select(driver.find_element(By.CSS_SELECTOR, 'select[id="str_shop_code"]'))
        store_select.select_by_visible_text(store_name)
        time.sleep(0.2)

        # 시작일
        start_date_input = driver.find_element(By.CSS_SELECTOR, 'input[id="start_date"]')
        start_date_input.clear()
        start_date_input.send_keys(self.guiDto.target_date)

        # 종료일
        end_date_input = driver.find_element(By.CSS_SELECTOR, 'input[id="end_date"]')
        end_date_input.clear()
        end_date_input.send_keys(self.guiDto.target_date)

        search_button = driver.find_element(By.XPATH, '//div[contains(@id, "search")][contains(text(), "검색")]')
        driver.execute_script("arguments[0].click();", search_button)
        time.sleep(3)

    def get_delivery_from_result(self, store_detail_dto: StoreDetailDto):
        driver = self.driver
        time.sleep(0.2)

        try:
            # 전체 결과 td
            # $x('//tr[./td[contains(text(), "2023-04-06")]]/td/a[contains(@href, "모두")]')
            delivery_result = driver.find_element(
                By.XPATH, f'//tr[./td[contains(text(), "{self.guiDto.target_date}")]]/td/a[contains(@href, "모두")]'
            ).get_attribute("textContent")
            store_detail_dto.delivery_result = delivery_result

        except Exception as e:
            print(f"검색 결과를 발견하지 못했습니다.")

        return store_detail_dto

    def get_discount_cost_from_store(self, store_detail_dto: StoreDetailDto):
        if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
            store_detail_dto = self.go_zigzag_and_search_discount_cost(store_detail_dto)

        elif store_detail_dto.store_name == StoreNameEnum.WeMakePrice.value:
            store_detail_dto = self.go_wemakeprice_and_search_discount_cost(store_detail_dto)

        elif store_detail_dto.store_name == StoreNameEnum.Coupang.value:
            store_detail_dto = self.go_coupang_and_search_discount_cost(store_detail_dto)

        elif store_detail_dto.store_name == StoreNameEnum.TicketMonster.value:
            print(self.dict_accounts["티몬"]["URL"])
            store_detail_dto = self.go_ticketmonster_and_search_discount_cost(store_detail_dto)

        else:
            return store_detail_dto

        return store_detail_dto

    def go_zigzag_and_search_discount_cost(self, store_detail_dto: StoreDetailDto):
        driver = self.driver
        login_url = self.dict_accounts["지그재그"]["URL"]

        try:
            # 새 탭에서 열기
            driver.execute_script(f"window.open('{login_url}');")
            driver.switch_to.window(driver.window_handles[1])

            self.zigzag_login()

            # 지그재그
            try:
                zigzag_cost = self.get_zigzag_cost()
            except Exception as e:
                print(e)
                print(f"지그재그 검색 실패")
                zigzag_cost = 0
            finally:
                store_detail_dto.zigzag_cost = zigzag_cost

            # 마이픽쿠폰
            try:
                mypick_cost = self.get_mypick_cost()
            except Exception as e:
                print(e)
                print(f"마이픽쿠폰 검색 실패")
                mypick_cost = 0
            finally:
                store_detail_dto.mypick_cost = mypick_cost

        except Exception as e:
            print(e)

        finally:
            # 원래 탭으로 돌아오기
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(0.5)
            return store_detail_dto

    def zigzag_login(self):
        driver = self.driver

        try:
            # 이전 로그인 세션이 남아있을 경우 바로 스토어 선택 화면으로 이동합니다.
            WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, '//h1[contains(text(), "파트너센터 로그인")]'))
            )
            time.sleep(0.2)

        except Exception as e:
            pass

        try:
            driver.implicitly_wait(1)

            login_id = self.dict_accounts["지그재그"]["ID"]
            login_pw = self.dict_accounts["지그재그"]["PW"]

            id_input = driver.find_element(By.XPATH, '//input[@placeholder="이메일"]')
            id_input.clear()
            time.sleep(0.2)
            id_input.send_keys(login_id)

            pw_input = driver.find_element(By.XPATH, '//input[@placeholder="비밀번호"]')
            pw_input.clear()
            time.sleep(0.2)
            pw_input.send_keys(login_pw)

            login_button = driver.find_element(By.XPATH, '//button[contains(text(), "로그인")]')
            login_button.click()
            time.sleep(0.2)

        except Exception as e:
            print("로그인 정보 입력 실패")

        finally:
            driver.implicitly_wait(self.default_wait)

        try:
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//div[contains(text(), "코코블랑")]'))
            )
            time.sleep(0.2)

            store_link = driver.find_element(By.XPATH, '//a[contains(@href, "cocoblanc")]')
            driver.execute_script("arguments[0].click();", store_link)
            time.sleep(0.2)

            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//span[contains(text(), "광고 관리")]'))
            )
            time.sleep(0.2)

        except Exception as e:
            print(e)
            raise Exception("지그재그 로그인 실패")

    def get_zigzag_cost(self):
        driver = self.driver
        search_date = self.guiDto.target_date.replace("-", "")
        driver.get(
            f"https://partners.kakaostyle.com/shop/cocoblanc/wallet?date_from_ymd={search_date}&date_to_ymd={search_date}&type=payment"
        )
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[contains(text(), "광고 코인 총 잔액")]'))
        )
        time.sleep(0.2)

        td_date = self.guiDto.target_date.replace("-", ".")
        result_trs = driver.find_elements(By.XPATH, f'//tr[./td[text()="{td_date}"]]')
        target_tr = result_trs[-1]

        zigzag_cost = (
            target_tr.find_element(By.CSS_SELECTOR, "td:nth-child(12)").get_attribute("textContent").replace("원", "")
        )

        return zigzag_cost

    def get_mypick_cost(self):
        driver = self.driver
        driver.get(f"https://partners.kakaostyle.com/shop/cocoblanc/coupon/my_pick/list")
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//h1[contains(text(), "마이픽쿠폰 관리")]'))
        )
        time.sleep(0.2)

        # $x('//table/tbody/tr')
        # 현재 검색 결과가 없음
        mypick_trs = driver.find_elements(By.XPATH, "//table/tbody/tr")

        mypick_cost = 0
        if len(mypick_trs) < 0:
            print("마이픽쿠폰 작업")
            mypick_cost = 0

        return mypick_cost

    def go_wemakeprice_and_search_discount_cost(self, store_detail_dto: StoreDetailDto):
        driver = self.driver
        login_url = self.dict_accounts["위메프"]["URL"]

        try:
            # 새 탭에서 열기
            driver.execute_script(f"window.open('{login_url}');")
            driver.switch_to.window(driver.window_handles[1])

            self.wemakeprice_login()

            # 파트너 부담 쿠폰 금액
            try:
                coupon_cost = self.wemakeprice_get_coupon_cost()
            except Exception as e:
                print(e)
                print(f"쿠폰 검색 실패")
                coupon_cost = 0
            finally:
                store_detail_dto.coupon_cost = coupon_cost

        except Exception as e:
            print(e)

        finally:
            # 원래 탭으로 돌아오기
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(0.5)
            return store_detail_dto

    def wemakeprice_login(self):
        driver = self.driver

        try:
            # 이전 로그인 세션이 남아있을 경우 바로 스토어 화면으로 이동합니다.
            WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//input[@name="loginid"]')))
            time.sleep(0.2)

        except Exception as e:
            pass

        try:
            driver.implicitly_wait(1)

            login_id = self.dict_accounts["위메프"]["ID"]
            login_pw = self.dict_accounts["위메프"]["PW"]

            id_input = driver.find_element(By.XPATH, '//input[@name="loginid"]')
            id_input.clear()
            time.sleep(0.2)
            id_input.send_keys(login_id)

            pw_input = driver.find_element(By.XPATH, '//input[@name="loginpassword"]')
            pw_input.clear()
            time.sleep(0.2)
            pw_input.send_keys(login_pw)

            login_button = driver.find_element(By.XPATH, '//button[contains(text(), "로그인")]')
            login_button.click()
            time.sleep(0.2)

        except Exception as e:
            print("로그인 정보 입력 실패")

        finally:
            driver.implicitly_wait(self.default_wait)

        try:
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//img[@alt="위메프 파트너 2.0"]')))
            time.sleep(0.2)

        except Exception as e:
            print(e)
            raise Exception("위메프 로그인 실패")

    def wemakeprice_get_coupon_cost(self):
        driver = self.driver
        driver.get(f"https://wpartner.wemakeprice.com/settle/dailySettleList")
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//h2[contains(text(), "매출현황")]')))
        time.sleep(0.2)

        # 기간 설정
        start_date_input = driver.find_element(By.XPATH, '//input[@id="startDt_D"]')
        start_date_input.clear()
        time.sleep(0.2)
        start_date_input.send_keys(self.guiDto.target_date)

        end_date_input = driver.find_element(By.XPATH, '//input[@id="endDt_D"]')
        end_date_input.clear()
        time.sleep(0.2)
        end_date_input.send_keys(self.guiDto.target_date)

        # 검색 클릭
        date_search_button = driver.find_element(By.XPATH, '//button[@id="searchBtn"]')
        date_search_button.click()
        time.sleep(1)

        coupon_cost = driver.find_element(By.XPATH, '//span[@id="settleCompleteTotSellerCouponAmt"]').get_attribute(
            "textContent"
        )
        coupon_cost = coupon_cost.replace(",", "")

        return coupon_cost

    def go_coupang_and_search_discount_cost(self, store_detail_dto: StoreDetailDto):
        driver = self.driver
        login_url = self.dict_accounts["쿠팡"]["URL"]

        try:
            # 새 탭에서 열기
            driver.execute_script(f"window.open('{login_url}');")
            driver.switch_to.window(driver.window_handles[1])

            self.coupang_login()

            # 쿠폰
            try:
                coupon_cost = self.coupang_get_coupon_cost()
            except Exception as e:
                print(e)
                print(f"쿠폰 검색 실패")
                coupon_cost = 0
            finally:
                store_detail_dto.coupon_cost = coupon_cost

        except Exception as e:
            print(e)

        finally:
            # 원래 탭으로 돌아오기
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(0.5)
            return store_detail_dto

    def coupang_login(self):
        driver = self.driver

        try:
            # 이전 로그인 세션이 남아있을 경우 바로 스토어 화면으로 이동합니다.
            WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, '//h1[contains(text(), "coupang")]'))
            )
            time.sleep(0.2)

        except Exception as e:
            pass

        try:
            driver.implicitly_wait(1)

            login_id = self.dict_accounts["쿠팡"]["ID"]
            login_pw = self.dict_accounts["쿠팡"]["PW"]

            id_input = driver.find_element(By.XPATH, '//input[@id="username"]')
            id_input.clear()
            time.sleep(0.2)
            id_input.send_keys(login_id)

            pw_input = driver.find_element(By.XPATH, '//input[@id="password"]')
            pw_input.clear()
            time.sleep(0.2)
            pw_input.send_keys(login_pw)

            login_button = driver.find_element(By.XPATH, '//input[contains(@id, "login")]')
            login_button.click()
            time.sleep(0.2)

        except Exception as e:
            print("로그인 정보 입력 실패")

        finally:
            driver.implicitly_wait(self.default_wait)

        try:
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//button[contains(text(), "Coupang Wing")]'))
            )
            time.sleep(0.2)

        except Exception as e:
            print(e)
            raise Exception("쿠팡 로그인 실패")

    def coupang_get_coupon_cost(self):
        driver = self.driver
        driver.get(
            f"https://wing.coupang.com/tenants/finance/wing/goods/purchase-report?from={self.guiDto.target_date}&to={self.guiDto.target_date}"
        )
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//h4[contains(text(), "매출내역")]')))
        time.sleep(0.2)

        coupon_cost = driver.find_element(By.XPATH, '//td[contains(text(), "합계:")]').get_attribute("textContent")
        coupon_cost = coupon_cost.replace("합계:", "")

        return coupon_cost

    def go_ticketmonster_and_search_discount_cost(self, store_detail_dto: StoreDetailDto):
        driver = self.driver
        login_url = self.dict_accounts["티몬"]["URL"]

        try:
            # 새 탭에서 열기
            driver.execute_script(f"window.open('{login_url}');")
            driver.switch_to.window(driver.window_handles[1])

            self.ticketmonster_login()

            # 쿠폰
            try:
                coupon_cost = self.ticketmonster_get_coupon_cost()
            except Exception as e:
                print(e)
                print(f"쿠폰 검색 실패")
                coupon_cost = 0
            finally:
                store_detail_dto.coupon_cost = coupon_cost

        except Exception as e:
            print(e)

        finally:
            # 원래 탭으로 돌아오기
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(0.5)
            return store_detail_dto

    def ticketmonster_login(self):
        driver = self.driver

        try:
            # 이전 로그인 세션이 남아있을 경우 바로 스토어 선택 화면으로 이동합니다.
            WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, '//h2[contains(text(), "파트너  로그인")]'))
            )
            time.sleep(0.2)

        except Exception as e:
            pass

        try:
            driver.implicitly_wait(1)

            login_id = self.dict_accounts["티몬"]["ID"]
            login_pw = self.dict_accounts["티몬"]["PW"]

            id_input = driver.find_element(By.XPATH, '//input[@id="form_id"]')
            id_input.clear()
            time.sleep(0.2)
            id_input.send_keys(login_id)

            pw_input = driver.find_element(By.XPATH, '//input[@id="form_password"]')
            pw_input.clear()
            time.sleep(0.2)
            pw_input.send_keys(login_pw)

            login_button = driver.find_element(By.XPATH, '//button[contains(@onclick, "submitLogin()")]')
            login_button.click()
            time.sleep(0.2)

            WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, '//button[contains(text(), "다음에 변경")]'))
            )
            time.sleep(0.2)

            change_next_time_button = driver.find_element(By.XPATH, '//button[contains(text(), "다음에 변경")]')
            change_next_time_button.click()
            time.sleep(0.2)

        except Exception as e:
            print("로그인 정보 입력 실패")

        finally:
            driver.implicitly_wait(self.default_wait)

        try:
            WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//h1[./a[contains(text(), "TMON 배송상품 파트너센터")]]'))
            )
            time.sleep(0.2)

        except Exception as e:
            print(e)
            raise Exception("티몬 로그인 실패")

    def ticketmonster_get_coupon_cost(self):
        driver = self.driver
        driver.get(f"https://spc-settlement.tmon.co.kr/revenue/sales")
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//h2[contains(text(), "판매현황 조회")]'))
        )
        time.sleep(0.2)

        # 할인현황
        driver.find_element(By.XPATH, '//button[contains(text(), "할인현황")]').click()
        time.sleep(0.2)

        # 기간 설정
        start_date_input = driver.find_element(By.XPATH, '//input[@id="startDate"]')
        start_date_input.clear()
        time.sleep(0.2)
        start_date_input.send_keys(self.guiDto.target_date)

        end_date_input = driver.find_element(By.XPATH, '//input[@id="endDate"]')
        end_date_input.clear()
        time.sleep(0.2)
        end_date_input.send_keys(self.guiDto.target_date)

        # 진행 딜 선택 '전체(전체)'
        deal_select = Select(driver.find_element(By.CSS_SELECTOR, 'select[name="mainDealSrl"]'))
        deal_select.select_by_visible_text("전체(전체)")
        time.sleep(0.2)

        # 검색 클릭
        date_search_button = driver.find_element(By.XPATH, '//button[@id="btn_srch"]')
        date_search_button.click()
        time.sleep(0.2)

        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//h3[contains(text(), "할인현황")]')))
        time.sleep(0.2)

        coupon_cost = driver.find_element(By.XPATH, '//tr[./td[contains(text(), "파트너부담 즉시할인 사용금액")]]//p').get_attribute(
            "textContent"
        )
        coupon_cost = coupon_cost.replace(",", "")

        return coupon_cost

    def update_excel_from_dto(self, target_date_row, store_min_col, store_detail_dto: StoreDetailDto):
        # 주문수량
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                sheet_coord = Cafe24Enum.주문수량.value
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                sheet_coord = ElevenStreetEnum.주문수량.value
            else:
                sheet_coord = CommonStoreEnum.주문수량.value

            tot_products_cell = self.sheet.cell(row=target_date_row, column=store_min_col + sheet_coord)
            original_value = tot_products_cell.value
            tot_products_cell.value = store_detail_dto.tot_products

        except Exception as e:
            print(e)
            tot_products_cell.value = original_value

        # 주문금액
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                sheet_coord = Cafe24Enum.주문금액.value
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                sheet_coord = ElevenStreetEnum.주문금액.value
            else:
                sheet_coord = CommonStoreEnum.주문금액.value

            tot_amount_cell = self.sheet.cell(row=target_date_row, column=store_min_col + sheet_coord)
            original_value = tot_amount_cell.value
            tot_amount_cell.value = store_detail_dto.tot_amount

        except Exception as e:
            print(e)
            tot_amount_cell.value = original_value

        # 원가금액
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                sheet_coord = Cafe24Enum.원가금액.value
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                sheet_coord = ElevenStreetEnum.원가금액.value
            else:
                sheet_coord = CommonStoreEnum.원가금액.value

            org_price_cell = self.sheet.cell(row=target_date_row, column=store_min_col + sheet_coord)
            original_value = org_price_cell.value
            org_price_cell.value = store_detail_dto.org_price

        except Exception as e:
            print(e)
            org_price_cell.value = original_value

        # 취소수량
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                sheet_coord = Cafe24Enum.취소수량.value
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                sheet_coord = ElevenStreetEnum.취소수량.value
            else:
                sheet_coord = CommonStoreEnum.취소수량.value

            cancel_total_data_product_sum_cell = self.sheet.cell(
                row=target_date_row, column=store_min_col + sheet_coord
            )
            original_value = cancel_total_data_product_sum_cell.value
            cancel_total_data_product_sum_cell.value = store_detail_dto.cancel_total_data_product_sum

        except Exception as e:
            print(e)
            cancel_total_data_product_sum_cell.value = original_value

        # 취소금액
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                raise Exception("카페24는 취소금액이 없습니다.")
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                sheet_coord = ElevenStreetEnum.취소금액.value
            else:
                sheet_coord = CommonStoreEnum.취소금액.value

            cancel_total_data_order_sum_amount_cell = self.sheet.cell(
                row=target_date_row, column=store_min_col + sheet_coord
            )
            original_value = cancel_total_data_order_sum_amount_cell.value
            cancel_total_data_order_sum_amount_cell.value = store_detail_dto.cancel_total_data_order_sum_amount

        except Exception as e:
            print(e)
            if str(e).find("없습니다") > -1:
                pass
            else:
                cancel_total_data_order_sum_amount_cell.value = original_value

        # 반품수량
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                sheet_coord = Cafe24Enum.반품수량.value
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                sheet_coord = ElevenStreetEnum.반품수량.value
            else:
                sheet_coord = CommonStoreEnum.반품수량.value

            refund_total_data_product_sum_cell = self.sheet.cell(
                row=target_date_row, column=store_min_col + sheet_coord
            )
            original_value = refund_total_data_product_sum_cell.value
            refund_total_data_product_sum_cell.value = store_detail_dto.refund_total_data_product_sum

        except Exception as e:
            print(e)
            refund_total_data_product_sum_cell.value = original_value

        # 반품금액
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                raise Exception("카페24는 반품금액이 없습니다.")
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                sheet_coord = ElevenStreetEnum.반품금액.value
            else:
                sheet_coord = CommonStoreEnum.반품금액.value

            refund_total_data_order_sum_amount_cell = self.sheet.cell(
                row=target_date_row, column=store_min_col + sheet_coord
            )
            original_value = refund_total_data_order_sum_amount_cell.value
            refund_total_data_order_sum_amount_cell.value = store_detail_dto.refund_total_data_order_sum_amount

        except Exception as e:
            print(e)
            if str(e).find("없습니다") > -1:
                pass
            else:
                refund_total_data_order_sum_amount_cell.value = original_value

        # 환불금액 = 취소금액 + 반품금액 (카페24만 기록)
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                sheet_coord = Cafe24Enum.환불금액.value
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                raise Exception("환불금액이 없습니다.")
            else:
                raise Exception("환불금액이 없습니다.")

            total_cancel_cell = self.sheet.cell(row=target_date_row, column=store_min_col + sheet_coord)
            original_value = total_cancel_cell.value
            total_cancel_cell.value = (
                store_detail_dto.refund_total_data_order_sum_amount
                + store_detail_dto.cancel_total_data_order_sum_amount
            )

        except Exception as e:
            print(e)
            if str(e).find("없습니다") > -1:
                pass
            else:
                total_cancel_cell.value = original_value

        # 지그재그 (카페24만 기록)
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                sheet_coord = Cafe24Enum.광고비지그재그.value
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                raise Exception("지그재그가 없습니다.")
            else:
                raise Exception("지그재그가 없습니다.")

            zigzag_cost_cell = self.sheet.cell(row=target_date_row, column=store_min_col + sheet_coord)
            original_value = zigzag_cost_cell.value
            zigzag_cost_cell.value = store_detail_dto.zigzag_cost

        except Exception as e:
            print(e)
            if str(e).find("없습니다") > -1:
                pass
            else:
                zigzag_cost_cell.value = original_value

        # 마이픽쿠폰 (카페24만 기록)
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                sheet_coord = Cafe24Enum.광고비마이픽쿠폰.value
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                raise Exception("마이픽쿠폰이 없습니다.")
            else:
                raise Exception("마이픽쿠폰이 없습니다.")

            mypick_cost_cell = self.sheet.cell(row=target_date_row, column=store_min_col + sheet_coord)
            original_value = mypick_cost_cell.value
            mypick_cost_cell.value = store_detail_dto.mypick_cost

        except Exception as e:
            print(e)
            if str(e).find("없습니다") > -1:
                pass
            else:
                mypick_cost_cell.value = original_value

        # 쿠폰비
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                raise Exception("쿠폰비가 없습니다.")
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                raise Exception("쿠폰비가 없습니다.")
            else:
                sheet_coord = CommonStoreEnum.쿠폰비.value

            coupon_cost_cell = self.sheet.cell(row=target_date_row, column=store_min_col + sheet_coord)
            original_value = coupon_cost_cell.value
            coupon_cost_cell.value = store_detail_dto.coupon_cost

        except Exception as e:
            print(e)
            if str(e).find("없습니다") > -1:
                pass
            else:
                coupon_cost_cell.value = original_value

        # 배송건수
        try:
            if store_detail_dto.store_name == StoreNameEnum.Cafe24.value:
                sheet_coord = Cafe24Enum.배송건수.value
            elif store_detail_dto.store_name == StoreNameEnum.ElevenStreet.value:
                sheet_coord = ElevenStreetEnum.배송건수.value
            else:
                sheet_coord = CommonStoreEnum.배송건수.value

            delivery_result_cell = self.sheet.cell(row=target_date_row, column=store_min_col + sheet_coord)
            original_value = delivery_result_cell.value
            delivery_result_cell.value = store_detail_dto.delivery_result

        except Exception as e:
            print(e)
            delivery_result_cell.value = original_value

        self.workbook.save(self.guiDto.stats_file)

        self.log_msg.emit(f"[{self.guiDto.target_date}] {store_detail_dto.store_name} 저장 완료")

        print(f"[{self.guiDto.target_date}] {store_detail_dto.store_name} 저장 완료")

    # 전체작업 시작
    def work_start(self):
        print(f"process: work_start")

        try:
            # 계정 엑셀 파일
            self.dict_accounts = self.get_dict_account()

            # 통계 엑셀 파일에서 상점 이름을 추출합니다.
            store_list = self.get_store_list()

            self.workbook = load_workbook(self.guiDto.stats_file)

            self.sheet = self.workbook[self.guiDto.sheet_name]

            target_date_row = self.get_target_date_row(self.guiDto.target_date)

            self.ezadmin_login()

            for store_name in store_list:
                print(f"{store_name} 작업 시작")

                self.log_msg.emit(f"{store_name} 작업 시작")

                store_detail_dto = StoreDetailDto()

                # # 스토어 테스트용 코드
                # if store_name != "티몬":
                #     continue

                try:
                    store_min_col = self.get_store_min_col(store_name)

                    self.go_store_calculate_menu_and_search_date()

                    store_detail_dto = self.get_calculate_from_result(store_name, store_detail_dto)

                    self.go_store_cancel_menu_and_search_date(store_detail_dto.store_name, "취소")

                    store_detail_dto = self.get_cancel_from_result(store_detail_dto)

                    self.go_store_cancel_menu_and_search_date(store_detail_dto.store_name, "반품")

                    store_detail_dto = self.get_refund_from_result(store_detail_dto)

                    self.go_store_delivery_menu_and_search_date(store_detail_dto.store_name)

                    store_detail_dto = self.get_delivery_from_result(store_detail_dto)

                    store_detail_dto = self.get_discount_cost_from_store(store_detail_dto)

                except Exception as e:
                    print(str(e))
                    self.log_msg.emit(f"{store_name} 작업 실패")
                    global_log_append(str(e))
                    continue

                finally:
                    self.update_excel_from_dto(target_date_row, store_min_col, store_detail_dto)

        except Exception as e:
            print(str(e))
            if str(e).find("채널명") > -1:
                self.log_msg.emit(f"계정 엑셀 파일 양식이 아닙니다.")
            else:
                self.log_msg.emit(f"{str(e)}")

        finally:
            self.driver.close()
            self.workbook.close()
            time.sleep(0.2)


if __name__ == "__main__":
    process = EzadminCrawlerProcess()
    process.work_start()
